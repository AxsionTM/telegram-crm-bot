from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.models.application import Application


DATA_DIR = Path("app/data")
EXCEL_FILE = DATA_DIR / "applications.xlsx"


class ExcelStorage:
    """Работа с Excel-файлом заявок."""

    HEADERS = [
        "ID",
        "Date",
        "Name",
        "Phone",
        "Description",
        "Status",
    ]

    def __init__(self):
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        self._create_if_not_exists()

    def _create_if_not_exists(self) -> None:
        if EXCEL_FILE.exists():
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Applications"
        sheet.append(self.HEADERS)
        workbook.save(EXCEL_FILE)

    def _load(self):
        return load_workbook(EXCEL_FILE)

    def get_next_id(self) -> int:
        workbook = self._load()
        sheet = workbook.active
        return sheet.max_row

    def create(
        self,
        name: str,
        phone: str,
        description: str,
    ) -> Application:
        workbook = self._load()
        sheet = workbook.active

        application = Application(
            id=self.get_next_id(),
            date=datetime.now().strftime("%d.%m.%Y %H:%M"),
            name=name,
            phone=phone,
            description=description,
            status="Новая",
        )

        sheet.append(
            [
                application.id,
                application.date,
                application.name,
                application.phone,
                application.description,
                application.status,
            ]
        )

        workbook.save(EXCEL_FILE)
        return application

    def get_all(self) -> list[Application]:
        workbook = self._load()
        sheet = workbook.active

        applications: list[Application] = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue

            applications.append(
                Application(
                    id=int(row[0]),
                    date=str(row[1]),
                    name=str(row[2]),
                    phone=str(row[3]),
                    description=str(row[4]),
                    status=str(row[5]),
                )
            )

        return applications

    def get_by_id(self, application_id: int) -> Application | None:
        for application in self.get_all():
            if application.id == application_id:
                return application
        return None

    def update_status(self, application_id: int, status: str) -> Application | None:
        workbook = self._load()
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2):
            if row[0].value is not None and int(row[0].value) == application_id:
                row[5].value = status
                workbook.save(EXCEL_FILE)
                return Application(
                    id=int(row[0].value),
                    date=str(row[1].value),
                    name=str(row[2].value),
                    phone=str(row[3].value),
                    description=str(row[4].value),
                    status=status,
                )

        return None

    def delete(self, application_id: int) -> bool:
        workbook = self._load()
        sheet = workbook.active

        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if row[0].value is not None and int(row[0].value) == application_id:
                sheet.delete_rows(idx)
                workbook.save(EXCEL_FILE)
                return True

        return False
